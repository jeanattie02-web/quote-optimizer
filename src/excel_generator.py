from io import BytesIO
from typing import Any, Dict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from src.models import QuoteInput


def generer_excel_devis(quote_input: QuoteInput, res: Dict[str, Any]) -> bytes:
    """Génère un classeur Excel synthétique avec mise en forme professionnelle."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse Financière"

    # Styles
    font_titre = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    fill_header = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    font_section = Font(name="Calibri", size=11, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # En-tête
    ws.merge_cells("A1:D1")
    cell_titre = ws["A1"]
    cell_titre.value = "SYNTHÈSE DU DEVIS - QUOTE OPTIMIZER"
    cell_titre.font = font_titre
    cell_titre.fill = fill_header
    cell_titre.alignment = align_center

    # Données d'entrée (Main-d'œuvre & Paramètres)
    ws.append([])
    ws.append(["Paramètres du projet", "Valeur"])
    ws["A3"].font = font_section
    ws["B3"].font = font_section

    ws.append(["Jours Développeur Junior (300 €/j)", quote_input.jours_junior])
    ws.append(["Jours Développeur Confirmé (500 €/j)", quote_input.jours_confirme])
    ws.append(["Jours Développeur Senior (800 €/j)", quote_input.jours_senior])
    ws.append(["Frais de déplacement (€)", quote_input.frais_deplacement])
    ws.append(["Marge cible demandée", f"{quote_input.marge_cible * 100:.1f} %"])

    # Résultats Financiers
    ws.append([])
    ws.append(["Indicateur Financier", "Montant (€)"])
    ws["A10"].font = font_section
    ws["B10"].font = font_section

    ws.append(["Coût Main-d'œuvre Brut", res["cout_humain_total"]])
    ws.append(["Frais Généraux (5%)", res["frais_generaux"]])
    ws.append(["Coût de Revient Total", res["cout_revient_total"]])
    ws.append(["Prix de Vente Conseillé", res["prix_vente_conseille"]])
    ws.append(["Marge Brute Réalisée", res["marge_brute_euros"]])

    # Mise en forme des colonnes
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    for row in range(11, 16):
        ws[f"B{row}"].number_format = "#,##0.00 €"
        ws[f"B{row}"].alignment = align_right

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
